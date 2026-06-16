import os
import json
import base64
import io
import anthropic
from http.server import BaseHTTPRequestHandler

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

SYSTEM_PROMPT = """Du er Fossekall, en AI-assistent som hjelper Blåfall AS med å skrive konsesjonssøknader til NVE.

Når du mottar informasjon fra brukeren, skal du skrive en kort og presis oppsummering av hva du har forstått.
Strukturer oppsummeringen slik:
- Prosjektnavn og lokasjon
- Type tiltak (elvekraftverk, pumpekraftverk, etc.)
- Nøkkeltall du har fått (effekt, produksjon, fallhøyde, etc.)
- Vedlagte dokumenter og hva de inneholder
- Hva som mangler eller er uklart

Skriv på norsk. Vær konkret og tydelig. Avslutt med å spørre om oppsummeringen stemmer."""


def extract_text(name, b64_content):
    if not b64_content:
        return None
    try:
        data = base64.b64decode(b64_content)
        ext = name.lower().rsplit('.', 1)[-1] if '.' in name else ''

        if ext == 'docx':
            from docx import Document
            doc = Document(io.BytesIO(data))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())

        elif ext == 'pdf':
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            return '\n'.join(page.extract_text() or '' for page in reader.pages)

        elif ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            for sheet in wb.worksheets:
                parts.append(f'[Ark: {sheet.title}]')
                for row in sheet.iter_rows(values_only=True):
                    line = '\t'.join(str(c) if c is not None else '' for c in row)
                    if line.strip():
                        parts.append(line)
            return '\n'.join(parts)

        elif ext in ('txt', 'md', 'csv'):
            return data.decode('utf-8', errors='replace')

    except Exception as e:
        return f"[Feil ved lesing av {name}: {e}]"

    return None


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length))

            user_message = body.get("message", "")
            files = body.get("files", [])

            content_parts = []
            if user_message:
                content_parts.append(user_message)

            for f in files:
                text = extract_text(f.get("name", ""), f.get("content"))
                if text:
                    content_parts.append(
                        f"\n--- Innhold fra {f['name']} ---\n{text[:40000]}"
                    )
                else:
                    content_parts.append(
                        f"\n[Vedlegg: {f.get('name', 'ukjent fil')} — kunne ikke lese innholdet]"
                    )

            content = '\n'.join(content_parts)

            message = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=1024,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}]
            )

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"summary": message.content[0].text}).encode())

        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
